import json
import os
import re
from pathlib import Path
from tempfile import NamedTemporaryFile

from flask import Flask, flash, render_template, request
from openpyxl import load_workbook
from pypdf import PdfReader


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
REFERENCE_PATH = DATA_DIR / "reference_data.json"

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-change-on-render")


MODEL_TIER_ORDER = [
    "F5 r2600",
    "F5 r2800",
    "F5 r4600",
    "F5 r4800",
    "F5 r5600",
    "F5 r5800",
    "F5 r5900",
    "F5 r10600",
    "F5 r10800",
    "F5 r10900",
    "F5 r12600-DS",
    "F5 r12800-DS",
    "F5 r12900-DS",
]
MODEL_TIERS = {model.lower(): index + 1 for index, model in enumerate(MODEL_TIER_ORDER)}

QUALITATIVE_FEATURES = {
    "redundant_power": {
        "label": "전원 이중화",
        "keywords": ["전원", "power", "psu", "dual power", "redundant power"],
    },
    "redundant_fan": {
        "label": "FAN 이중화/Hot-Swap",
        "keywords": ["fan", "팬", "hot-swap", "hotswap", "hot swap"],
    },
}


@app.after_request
def add_no_store_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def load_reference():
    if not REFERENCE_PATH.exists():
        return {"f5_models": [], "comparisons": [], "sources": [], "pdf_summary": {}}
    return json.loads(REFERENCE_PATH.read_text(encoding="utf-8"))


def text_from_upload(uploaded_file):
    filename = uploaded_file.filename or "uploaded"
    suffix = Path(filename).suffix.lower()

    if suffix in {".txt", ".csv"}:
        raw = uploaded_file.read()
        return raw.decode("utf-8-sig", errors="ignore")

    with NamedTemporaryFile(delete=False, suffix=suffix) as temp:
        uploaded_file.save(temp.name)
        temp_path = Path(temp.name)

    try:
        if suffix == ".pdf":
            reader = PdfReader(str(temp_path))
            return "\n".join(page.extract_text() or "" for page in reader.pages)

        if suffix in {".xlsx", ".xlsm"}:
            workbook = load_workbook(str(temp_path), data_only=True)
            chunks = []
            for sheet in workbook.worksheets:
                chunks.append(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    values = [clean_text(cell) for cell in row if clean_text(cell)]
                    if values:
                        chunks.append(" | ".join(values))
            return "\n".join(chunks)

        raw = uploaded_file.read()
        return raw.decode("utf-8-sig", errors="ignore")
    finally:
        try:
            temp_path.unlink()
        except OSError:
            pass


def number_with_unit(value):
    if value is None:
        return 0.0
    text = str(value).replace(",", "").strip().lower()
    found = re.search(r"(\d+(?:\.\d+)?)\s*(gb|g|k|m)?\b", text)
    if not found:
        return 0.0
    number = float(found.group(1))
    unit = found.group(2) or ""
    if unit == "k":
        return number * 1_000
    if unit == "m":
        return number * 1_000_000
    return number


def parse_requested_number(text):
    normalized = str(text or "").replace(",", "").strip().lower()
    found = re.search(r"(\d+(?:\.\d+)?)\s*(억|만|gbps|ge|gb|g|tps|cps|k|m)?", normalized)
    if not found:
        return None
    number = float(found.group(1))
    unit = found.group(2) or ""
    if unit == "억":
        return number * 100_000_000
    if unit == "만":
        return number * 10_000
    if unit == "k":
        return number * 1_000
    if unit == "m":
        return number * 1_000_000
    return number


def parse_throughput_pair(value):
    text = str(value or "")
    numbers = [float(item) for item in re.findall(r"(\d+(?:\.\d+)?)\s*G", text, re.IGNORECASE)]
    if len(numbers) >= 2:
        return numbers[0], numbers[1]
    if len(numbers) == 1:
        return numbers[0], numbers[0]
    return 0.0, 0.0


def parse_ports(interface_text):
    text = str(interface_text or "").lower()
    ports = {
        "sfp_1g_ports": 0,
        "sfp_plus_10g_ports": 0,
        "sfp28_25g_ports": 0,
        "copper_1g_ports": 0,
        "copper_10g_ports": 0,
        "qsfp_40g_ports": 0,
        "qsfp28_100g_ports": 0,
    }

    for count, desc in re.findall(r"(\d+)\s*x\s*([^,]+)", text):
        qty = int(count)
        is_copper = any(token in desc for token in ["copper", "utp", "rj45"])
        is_sfp28 = "sfp28" in desc
        is_sfp_plus = "sfp+" in desc or "sfp" in desc
        is_qsfp = "qsfp" in desc

        if is_copper:
            if "1g" in desc:
                ports["copper_1g_ports"] += qty
            if "10g" in desc:
                ports["copper_10g_ports"] += qty
            continue

        if is_sfp28 and "25g" in desc:
            ports["sfp28_25g_ports"] += qty
        if is_sfp_plus and "10g" in desc:
            ports["sfp_plus_10g_ports"] += qty
        if is_sfp_plus and "1g" in desc:
            ports["sfp_1g_ports"] += qty
        if is_qsfp and "40g" in desc:
            ports["qsfp_40g_ports"] += qty
        if is_qsfp and "100g" in desc:
            ports["qsfp28_100g_ports"] += qty

    return ports


def display_model_name(model_name):
    name = str(model_name or "").replace("F5 ", "F5 BIG-IP ")
    return re.sub(r"\br(\d)", r"R\1", name)


def tier_for_model(model_name):
    return MODEL_TIERS.get(str(model_name or "").lower(), 999)


def normalize_model(model):
    specs = model.get("specs", {})
    l4, l7 = parse_throughput_pair(specs.get("L4/L7 Throughput"))
    model_name = model.get("model", "")
    ports = parse_ports(specs.get("인터페이스"))
    tier = int(model.get("tier") or tier_for_model(model_name))

    return {
        "model": model_name,
        "display_model": display_model_name(model_name),
        "tier": tier,
        "series": model.get("series", ""),
        "raw_specs": specs,
        "l4_gbps": l4,
        "l7_gbps": l7,
        "l4_cps": number_with_unit(specs.get("L4 CPS")),
        "ssl_tps": number_with_unit(
            specs.get("SSL TPS (RSA 2K)") or specs.get("SSL TPS RSA 2K")
        ),
        "concurrent_connections": number_with_unit(specs.get("동시 커넥션")),
        "ssd_gb": number_with_unit(specs.get("Storage")),
        "memory_gb": number_with_unit(specs.get("Memory")),
        **ports,
        "redundant_power": True,
        "redundant_fan": True,
    }


def extract_threshold(text, keywords, allowed_units=None, require_condition=True):
    lowered = text.lower()
    for keyword in keywords:
        key = keyword.lower()
        idx = lowered.find(key)
        if idx < 0:
            continue
        window = lowered[idx + len(key) : idx + len(key) + 100]
        if require_condition and not any(token in window for token in ["이상", "최소", "지원", ">=", "more", "over"]):
            continue
        found = re.search(r"(\d+(?:,\d{3})*(?:\.\d+)?)\s*(억|만|gbps|ge|gb|g|tps|cps|k|m)?", window)
        if found:
            unit = (found.group(2) or "").lower()
            if allowed_units is not None and unit not in allowed_units:
                continue
            return parse_requested_number(found.group(0))
    return None


def extract_port_requirement(text, keywords):
    lowered = text.lower()
    for keyword in keywords:
        idx = lowered.find(keyword.lower())
        if idx < 0:
            continue
        window = lowered[max(0, idx - 50) : idx + 90]
        for pattern in [r"(\d+)\s*(?:port|ports|포트)", r"(\d+)\s*(?:개|ea)\s*(?:이상)?"]:
            found = re.search(pattern, window, re.IGNORECASE)
            if found:
                return int(found.group(1))
    return None


def parse_requirements(text):
    requirements = []
    numeric_checks = [
        (
            "l4_gbps",
            "L4 Throughput",
            ["l4 throughput", "l4 처리량", "처리 성능", "throughput"],
            "Gbps",
            {"gbps", "g"},
        ),
        (
            "l7_gbps",
            "L7 Throughput",
            ["l7 throughput", "l7 처리량"],
            "Gbps",
            {"gbps", "g"},
        ),
        ("l4_cps", "L4 CPS", ["l4 cps", "connections per second", "cps"], "CPS", {"cps", "만", "억", "k", "m"}),
        ("ssl_tps", "SSL TPS", ["ssl 2k tps", "ssl tps"], "TPS", {"tps", "만", "억", "k", "m"}),
        (
            "concurrent_connections",
            "Concurrent Connection",
            ["concurrent connection", "동시 커넥션", "동시 세션", "동시접속", "동시 연결", "session"],
            "",
            {"만", "억", "k", "m"},
        ),
        ("ssd_gb", "SSD", ["ssd", "storage", "disk", "스토리지"], "GB", {"gb", "g"}),
        ("memory_gb", "Memory", ["memory", "메모리", "ram"], "GB", {"gb", "g"}),
    ]

    for key, label, keywords, unit, allowed_units in numeric_checks:
        value = extract_threshold(text, keywords, allowed_units)
        if value is not None:
            requirements.append({"key": key, "label": label, "value": value, "unit": unit, "type": "numeric"})

    port_checks = [
        ("sfp28_25g_ports", "25G SFP28 Port", ["sfp28", "25g"]),
        ("sfp_plus_10g_ports", "10G SFP+ Port", ["10g sfp+", "sfp+"]),
        ("sfp_1g_ports", "1G SFP Port", ["1g sfp"]),
        ("copper_10g_ports", "10G Copper/UTP/RJ45 Port", ["10g copper", "10g utp", "10g rj45"]),
        ("copper_1g_ports", "1G Copper/UTP/RJ45 Port", ["1g copper", "1g utp", "1g rj45", "utp", "copper", "rj45"]),
        ("qsfp28_100g_ports", "100G QSFP28 Port", ["100ge", "100g", "qsfp28"]),
        ("qsfp_40g_ports", "40G QSFP Port", ["40ge", "40g", "qsfp"]),
    ]

    for key, label, keywords in port_checks:
        if key == "copper_1g_ports" and any(
            token in text.lower() for token in ["10g copper", "10g utp", "10g rj45"]
        ):
            continue
        value = extract_port_requirement(text, keywords)
        if value is not None:
            requirements.append({"key": key, "label": label, "value": value, "unit": "Port", "type": "numeric"})

    lowered = text.lower()
    if any(token in lowered for token in QUALITATIVE_FEATURES["redundant_power"]["keywords"]) and (
        "이중화" in lowered or "dual" in lowered or "redundant" in lowered
    ):
        requirements.append({"key": "redundant_power", "label": "전원 이중화", "value": True, "unit": "", "type": "boolean"})

    if any(token in lowered for token in QUALITATIVE_FEATURES["redundant_fan"]["keywords"]) and (
        "이중화" in lowered or "hot" in lowered or "dual" in lowered or "redundant" in lowered
    ):
        requirements.append({"key": "redundant_fan", "label": "FAN 이중화/Hot-Swap", "value": True, "unit": "", "type": "boolean"})

    deduped = []
    seen = set()
    for req in requirements:
        if req["key"] in seen:
            continue
        seen.add(req["key"])
        deduped.append(req)
    return deduped


def format_requirement_value(req):
    value = req["value"]
    if value is True:
        return "필수"
    if req["key"] == "concurrent_connections":
        if value >= 1_000_000_000:
            return f"{value / 1_000_000_000:g}B 이상"
        return f"{value / 1_000_000:g}M 이상"
    if req["key"] == "ssl_tps":
        return f"{value:,.0f} TPS 이상"
    if req["key"] == "l4_cps":
        return f"{value:,.0f} CPS 이상"
    return f"{value:g}{req['unit']} 이상"


def format_model_value(model, key):
    value = model.get(key)
    if key == "concurrent_connections":
        if value >= 1_000_000_000:
            return f"{value / 1_000_000_000:g}B"
        return f"{value / 1_000_000:g}M"
    if key == "ssl_tps":
        return f"{value:,.0f} TPS"
    if key == "l4_cps":
        return f"{value:,.0f} CPS"
    if key in {"l4_gbps", "l7_gbps"}:
        return f"{value:g}Gbps"
    if key in {"ssd_gb", "memory_gb"}:
        return f"{value:g}GB"
    if key.endswith("_ports"):
        return f"{int(value or 0)}Port"
    if key in {"redundant_power", "redundant_fan"}:
        return "지원" if value else "확인 필요"
    return str(value)


def check_status(actual, required):
    if required is True:
        if actual is True:
            return "충족", "✅ 충족", "요구 조건 지원"
        return "확인 필요", "⚠️ 확인 필요", "자료상 지원 여부 확인 필요"

    if actual is None:
        return "확인 필요", "⚠️ 확인 필요", "장비 스펙 데이터 없음"
    if float(actual) < float(required):
        return "미충족", "❌ 미충족", "요구 기준보다 낮음"
    if float(required) > 0 and float(actual) >= float(required) * 1.2:
        return "여유 있음", "✅ 여유 있음", "요구 기준 대비 20% 이상 여유"
    return "충족", "✅ 충족", "요구 기준 충족"


def evaluate_model(model, requirements):
    checks = []
    for req in requirements:
        actual = model.get(req["key"])
        status, result_label, note = check_status(actual, req["value"])
        passed = status in {"충족", "여유 있음"}
        checks.append(
            {
                "key": req["key"],
                "label": req["label"],
                "required": format_requirement_value(req),
                "actual": format_model_value(model, req["key"]),
                "status": status,
                "result": result_label,
                "note": note,
                "passed": passed,
            }
        )

    total = len(checks)
    passed_count = sum(1 for check in checks if check["passed"])
    failed_checks = [check for check in checks if not check["passed"]]
    fit_score = round((passed_count / total) * 100, 1) if total else 0
    return {
        "model": model,
        "checks": checks,
        "passed_count": passed_count,
        "failed_count": total - passed_count,
        "failed_checks": failed_checks,
        "fit_score": fit_score,
        "all_passed": total > 0 and passed_count == total,
    }


def comparison_for_model(model_name, reference):
    for comparison in reference.get("comparisons", []):
        if comparison.get("f5_model") == model_name:
            return comparison
    return None


def metric_value(comparison, metric_name, product_name):
    if not comparison:
        return "-"
    for row in comparison.get("rows", []):
        if row.get("metric") == metric_name:
            return row.get("values", {}).get(product_name, "-") or "-"
    return "-"


def competitor_summary(model_name, reference):
    comparison = comparison_for_model(model_name, reference)
    if not comparison:
        return []

    products = [comparison.get("f5_model", "")] + comparison.get("competitors", [])
    return [
        {
            "product": product,
            "l4_l7": metric_value(comparison, "L4/L7 Throughput", product),
            "l4_cps": metric_value(comparison, "L4 CPS", product),
            "ssl_tps": metric_value(comparison, "SSL TPS (RSA 2K)", product)
            if metric_value(comparison, "SSL TPS (RSA 2K)", product) != "-"
            else metric_value(comparison, "SSL TPS RSA 2K", product),
            "connections": metric_value(comparison, "동시 커넥션", product),
        }
        for product in products
        if product
    ]


def recommend_model(requirements, reference):
    models = [normalize_model(model) for model in reference.get("f5_models", [])]
    evaluations = [evaluate_model(model, requirements) for model in models]
    for evaluation in evaluations:
        evaluation["competitors"] = competitor_summary(evaluation["model"]["model"], reference)

    matched = [evaluation for evaluation in evaluations if evaluation["all_passed"]]
    matched.sort(key=lambda item: item["model"]["tier"])
    alternatives = sorted(
        evaluations,
        key=lambda item: (-item["fit_score"], item["failed_count"], item["model"]["tier"]),
    )[:3]
    return (matched[0] if matched else None), alternatives, evaluations


def summarize_passed(evaluation, limit=5):
    if not evaluation:
        return []
    return [check["label"] for check in evaluation["checks"] if check["passed"]][:limit]


def mail_texts(recommendation, alternatives):
    primary = recommendation or (alternatives[0] if alternatives else None)
    model = primary["model"]["display_model"] if primary else "확인 필요"
    passed = summarize_passed(primary)
    passed_text = ", ".join(passed) if passed else "충족 항목 없음"
    alternative_names = [item["model"]["display_model"] for item in alternatives if item is not primary][:2]
    alternative_text = ", ".join(alternative_names) if alternative_names else "대안 모델 없음"

    if recommendation:
        simple = (
            f"문의 주신 스펙 기준으로는 {model} 모델이 적합합니다.\n"
            f"주요 충족 조건은 {passed_text}이며, 대안 모델은 {alternative_text}입니다."
        )
        detail = (
            f"고객 요구 스펙을 등록된 F5 rSeries 기준 데이터와 비교한 결과, {model} 모델이 모든 필수 조건을 충족합니다.\n"
            f"주요 충족 조건: {passed_text}\n"
            f"대안 검토 모델: {alternative_text}\n"
            "해당 모델 기준으로 제안 진행이 가능합니다."
        )
        internal = (
            f"[내부 검토]\n추천 모델: {model}\n적합도: {primary['fit_score']}%\n"
            f"주요 충족 조건: {passed_text}\n대안 모델: {alternative_text}\n"
            "검토 의견: 필수 조건을 모두 충족하는 최저 tier 모델입니다."
        )
    else:
        failed = [check["label"] for check in primary["failed_checks"]] if primary else []
        failed_text = ", ".join(failed) if failed else "확인 필요"
        simple = (
            "조건을 모두 만족하는 모델이 없습니다.\n"
            f"가장 근접한 모델은 {model}이며, 미충족 항목은 {failed_text}입니다."
        )
        detail = (
            "등록된 F5 rSeries 기준 데이터와 비교한 결과, 고객 요구 조건을 모두 만족하는 단일 모델은 없습니다.\n"
            f"가장 근접한 모델: {model}\n"
            f"주요 충족 조건: {passed_text}\n"
            f"미충족 조건: {failed_text}\n"
            f"대안 검토 모델: {alternative_text}\n"
            "상위 구성, 복수 장비 구성, 또는 요구 조건 조정 여부를 추가 확인해야 합니다."
        )
        internal = (
            f"[내부 검토]\n결론: 조건을 모두 만족하는 모델 없음\n"
            f"최근접 모델: {model}\n적합도: {primary['fit_score'] if primary else 0}%\n"
            f"미충족 조건: {failed_text}\n대안 모델: {alternative_text}\n"
            "검토 의견: 최종 제안 전 공식 Datasheet 및 구성 가능 여부 확인 필요."
        )

    return {"simple": simple, "detail": detail, "internal": internal}


@app.route("/", methods=["GET", "POST"])
def index():
    reference = load_reference()
    context = {
        "models": [normalize_model(model) for model in reference.get("f5_models", [])],
        "example_text": (
            "L4 Throughput 20Gbps 이상\n"
            "L7 Throughput 13Gbps 이상\n"
            "Concurrent Connection 19M 이상\n"
            "SSL TPS 7000 이상\n"
            "25G SFP28 4Port 이상\n"
            "10G Copper 4Port 이상\n"
            "SSD 480GB 이상\n"
            "전원/FAN 이중화"
        ),
    }

    if request.method == "POST":
        requirement_text = clean_text(request.form.get("requirement_text"))
        upload = request.files.get("requirement_file")
        if upload and upload.filename:
            uploaded_text = text_from_upload(upload)
            requirement_text = f"{requirement_text}\n{uploaded_text}".strip()

        if not requirement_text:
            flash("고객 요구 스펙을 입력하거나 파일을 업로드해주세요.")
            return render_template("index.html", **context)

        requirements = parse_requirements(requirement_text)
        if not requirements:
            flash("하드웨어 비교 조건을 찾지 못했습니다. L4/L7 Throughput, SSL TPS, CPS, 동시 세션, 포트, SSD처럼 수치와 단위가 있는 조건을 입력해주세요.")
            return render_template("index.html", requirement_text=requirement_text, **context)

        recommendation, alternatives, evaluations = recommend_model(requirements, reference)
        return render_template(
            "index.html",
            requirement_text=requirement_text,
            requirements=requirements,
            recommendation=recommendation,
            alternatives=alternatives,
            evaluations=evaluations,
            mail_texts=mail_texts(recommendation, alternatives),
            **context,
        )

    return render_template("index.html", **context)


if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=int(os.getenv("PORT", "5000")),
        debug=os.getenv("FLASK_DEBUG") == "1",
    )
