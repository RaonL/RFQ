import json
import os
from pathlib import Path

import openpyxl
from pypdf import PdfReader


BASE_DIR = Path(__file__).resolve().parent
PDF_PATH = Path(
    os.getenv("F5_DATA_SHEET_PATH", BASE_DIR / "source" / "F5_rSeries_Appliance_Data_Sheet.pdf")
)
XLSX_PATH = Path(
    os.getenv("F5_COMPARISON_XLSX_PATH", BASE_DIR / "source" / "r-Series_경쟁사_비교자료(20260504).xlsx")
)


def main() -> None:
    reader = PdfReader(str(PDF_PATH))
    preview_text = "\n".join((page.extract_text() or "") for page in reader.pages[:3])

    workbook = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        preview_rows = []
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(sheet.max_row, 10),
            values_only=True,
        ):
            preview_rows.append([str(cell) if cell is not None else "" for cell in row[:14]])

        sheets.append(
            {
                "sheet": sheet.title,
                "rows": sheet.max_row,
                "cols": sheet.max_column,
                "preview": preview_rows,
            }
        )

    print(
        json.dumps(
            {
                "pdf_pages": len(reader.pages),
                "pdf_preview": preview_text[:3000],
                "sheets": sheets,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
